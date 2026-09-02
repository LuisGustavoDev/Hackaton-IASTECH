"""
Relatório comparativo em Excel: um arquivo com tudo que se sabe sobre o
desempenho do detector.

Feito para ser rodado de novo a cada checkpoint novo. Ele agrupa as
execuções gravadas no banco POR CHECKPOINT, de modo que rodar o lote com
o modelo de 15 épocas e depois com o de 250 e gerar o relatório uma vez
produz as duas rodadas lado a lado, na mesma planilha.

ABAS
----
    Leia-me     como ler cada número e o que mais épocas podem ou não
                resolver
    Resumo      uma linha por checkpoint: VP/FP/FN, precisão, revocação,
                F1, tempos, e os metadados do treino (épocas, mAP, loss)
    Curva       precisão x revocação em cada limiar de score
    Por classe  desempenho por classe CRUZADO com quantos exemplos de
                treino a classe tem — é onde se vê o que é problema de
                época e o que é problema de dado
    Por imagem  quais diagramas o modelo pior atende
    Confusões   pares de classe trocados
    Detecções   uma linha por predição e por anotação perdida, com IoU e
                situação — a aba de análise manual

USO
---
    python -m app.diagnostico.relatorio --saida data/output/relatorio.xlsx
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.diagnostico import gabarito
from app.diagnostico.matriz_confusao import SEM_PAR, casar_detalhado
from app.diagnostico.utf8 import forcar_utf8
from app.models import execucoes

LIMIARES_CURVA = (0.05, 0.1, 0.15, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8)

NEGRITO = Font(bold=True)


# ======================================================================
# Coleta
# ======================================================================


def exemplos_de_treino(raiz: str | Path) -> Counter:
    """
    Quantas vezes cada classe aparece no conjunto de TREINO.

    É a coluna que transforma a tabela por classe em diagnóstico: uma
    classe com 3 exemplos e revocação zero não está esperando mais
    épocas, está esperando mais anotação.
    """
    contagem: Counter = Counter()

    for xml_path in Path(raiz).rglob("*.xml"):
        for anotacao in gabarito.caixas(xml_path):
            contagem[anotacao["classe"]] += 1

    return contagem


def metadados_checkpoint(caminho: str) -> dict:
    """
    Épocas, mAP e loss gravados no checkpoint, quando o arquivo existe.

    Best-effort de propósito: o relatório precisa funcionar numa máquina
    que tem o banco mas não tem mais o .pt daquela rodada.
    """
    arquivo = Path(caminho)

    if not arquivo.is_file():
        return {}

    try:
        from app.detection.checkpoint import carregar_checkpoint

        return carregar_checkpoint(arquivo).get("metadados", {}) or {}
    except Exception:
        return {}


def avaliar(execucao_ids, indice_gabarito, limiar_iou, limiar_score):
    """Cruza as execuções de uma rodada com o gabarito."""
    detalhes = []
    por_imagem = []
    matriz: Counter = Counter()
    tempos = []

    for execucao_id in execucao_ids:
        resumo = execucoes.obter(execucao_id)

        if resumo is None:
            continue

        xml_path = indice_gabarito.get(resumo.arquivo_nome or "")

        if xml_path is None:
            continue

        predicoes = [
            p
            for p in execucoes.deteccoes_de(execucao_id)
            if p["score"] >= limiar_score
        ]
        anotacoes = gabarito.caixas(xml_path)

        pares, fp, fn = casar_detalhado(predicoes, anotacoes, limiar_iou)

        vp_imagem = 0
        trocas_imagem = 0

        for anotacao, predicao, iou in pares:
            acertou = anotacao["classe"] == predicao["classe"]

            vp_imagem += acertou
            trocas_imagem += not acertou

            matriz[(anotacao["classe"], predicao["classe"])] += 1

            detalhes.append(
                _detalhe(
                    resumo,
                    "VP" if acertou else "CLASSE TROCADA",
                    anotacao["classe"],
                    predicao,
                    iou,
                )
            )

        for predicao in fp:
            matriz[(SEM_PAR, predicao["classe"])] += 1
            detalhes.append(
                _detalhe(resumo, "FP", SEM_PAR, predicao, 0.0)
            )

        for anotacao in fn:
            matriz[(anotacao["classe"], SEM_PAR)] += 1
            detalhes.append(
                _detalhe(resumo, "FN", anotacao["classe"], None, 0.0, anotacao)
            )

        por_imagem.append(
            {
                "execucao": execucao_id,
                "arquivo": resumo.arquivo_nome,
                "dimensoes": f"{resumo.imagem_largura}x{resumo.imagem_altura}",
                "anotadas": len(anotacoes),
                "previstas": len(predicoes),
                "vp": vp_imagem,
                "trocas": trocas_imagem,
                "fp": len(fp),
                "fn": len(fn),
                "revocacao": vp_imagem / len(anotacoes) if anotacoes else 0.0,
                "tags_lidas": resumo.qtd_tags_lidas,
                "tempo_total_ms": resumo.tempo_total_ms,
            }
        )

        if resumo.tempo_total_ms:
            tempos.append(resumo.tempo_total_ms)

    return matriz, por_imagem, detalhes, tempos


def _detalhe(resumo, situacao, classe_real, predicao, iou, anotacao=None):
    caixa = predicao or anotacao or {}

    return {
        "arquivo": resumo.arquivo_nome,
        "situacao": situacao,
        "classe_real": classe_real,
        "classe_prevista": predicao["classe"] if predicao else "",
        "score": round(predicao["score"], 4) if predicao else None,
        "iou": round(iou, 3) if iou else None,
        "x1": int(caixa.get("x1", 0)),
        "y1": int(caixa.get("y1", 0)),
        "x2": int(caixa.get("x2", 0)),
        "y2": int(caixa.get("y2", 0)),
        "tag": (predicao or {}).get("tag", ""),
        "texto_ocr": (predicao or {}).get("texto_bruto", ""),
    }


def totais(matriz: Counter) -> dict:
    vp = sum(q for (r, p), q in matriz.items() if r == p and r != SEM_PAR)
    fp = sum(q for (r, _), q in matriz.items() if r == SEM_PAR)
    fn = sum(q for (_, p), q in matriz.items() if p == SEM_PAR)
    trocas = sum(
        q
        for (r, p), q in matriz.items()
        if r != p and r != SEM_PAR and p != SEM_PAR
    )

    anotadas = vp + fn + trocas
    previstas = vp + fp + trocas

    precisao = vp / previstas if previstas else 0.0
    revocacao = vp / anotadas if anotadas else 0.0

    return {
        "vp": vp,
        "fp": fp,
        "fn": fn,
        "trocas": trocas,
        "anotadas": anotadas,
        "previstas": previstas,
        "precisao": precisao,
        "revocacao": revocacao,
        "f1": (
            2 * precisao * revocacao / (precisao + revocacao)
            if (precisao + revocacao)
            else 0.0
        ),
        "localizadas": (vp + trocas) / anotadas if anotadas else 0.0,
    }


# ======================================================================
# Escrita
# ======================================================================


def escrever(aba, cabecalho, linhas, larguras=None):
    aba.append(cabecalho)

    for celula in aba[1]:
        celula.font = NEGRITO
        celula.alignment = Alignment(horizontal="center", wrap_text=True)

    for linha in linhas:
        aba.append(linha)

    for indice, coluna in enumerate(cabecalho, start=1):
        letra = get_column_letter(indice)
        aba.column_dimensions[letra].width = (
            larguras[indice - 1] if larguras else max(len(str(coluna)) + 4, 11)
        )

    aba.freeze_panes = "A2"

    if linhas:
        aba.auto_filter.ref = (
            f"A1:{get_column_letter(len(cabecalho))}{len(linhas) + 1}"
        )


def aba_leiame(planilha, rodadas, limiar_iou, limiar_score):
    aba = planilha.active
    aba.title = "Leia-me"

    texto = [
        ("RELATÓRIO COMPARATIVO DO DETECTOR", True),
        ("", False),
        (f"Limiar de IoU usado no casamento: {limiar_iou}", False),
        (f"Limiar de score do relatório: {limiar_score}", False),
        (f"Rodadas comparadas: {len(rodadas)}", False),
        ("", False),
        ("COMO CADA NÚMERO É APURADO", True),
        (
            "Cada predição reivindica a anotação de maior IoU ainda livre, "
            "da mais confiante para a menos confiante. Depois disso:",
            False,
        ),
        ("  VP             caixa casada e classe certa", False),
        (
            "  CLASSE TROCADA caixa casada, rótulo errado — o modelo ACHOU "
            "o equipamento",
            False,
        ),
        ("  FP             previu onde não havia nada", False),
        ("  FN             anotação que nenhuma predição cobriu", False),
        ("", False),
        (
            "VN não aparece: detecção de objetos não tem verdadeiro "
            "negativo natural. Não existe um conjunto finito de caixas que "
            "poderiam ter sido previstas e corretamente não foram. Definir "
            "um VN exige uma regra de negócio a combinar com o cliente.",
            False,
        ),
        ("", False),
        ("O QUE MAIS ÉPOCAS PODEM RESOLVER", True),
        (
            "Compare a aba 'Por classe': a coluna 'exemplos no treino' ao "
            "lado da revocação separa dois problemas diferentes que se "
            "parecem no resultado.",
            False,
        ),
        (
            "  Muitos exemplos + revocação baixa  -> subtreino. Mais épocas "
            "tendem a ajudar.",
            False,
        ),
        (
            "  Poucos exemplos + revocação zero   -> falta de dado. Épocas "
            "não criam exemplo; tende a piorar por overfit.",
            False,
        ),
        ("", False),
        (
            "Olhe também 'CLASSE TROCADA' no Resumo. Ela mede equipamento "
            "encontrado com rótulo errado — é erro de classificação, não de "
            "detecção, e responde a época e a consistência da anotação, não "
            "a limiar.",
            False,
        ),
        ("", False),
        (
            "Na aba 'Confusões', um par que domina os demais costuma "
            "indicar fronteira mal definida entre duas classes no próprio "
            "gabarito. Vale reanotar antes de treinar de novo: nesse caso "
            "nenhuma quantidade de épocas conserta.",
            False,
        ),
        ("", False),
        ("PARA A ANÁLISE MANUAL", True),
        (
            "'Por imagem' ordenada pela pior revocação mostra onde começar "
            "a olhar. 'Detecções' traz uma linha por predição e por "
            "anotação perdida, com coordenadas — dá para abrir o diagrama "
            "ao lado e conferir caso a caso. Ambas têm filtro ligado.",
            False,
        ),
        ("", False),
        (
            "Para ver as caixas desenhadas sobre uma imagem: "
            "python -m app.detection.verificar --imagem <arquivo> --salvar saida.png",
            False,
        ),
    ]

    for indice, (linha, negrito) in enumerate(texto, start=1):
        celula = aba.cell(row=indice, column=1, value=linha)
        celula.font = NEGRITO if negrito else Font()
        celula.alignment = Alignment(wrap_text=True, vertical="top")

    aba.column_dimensions["A"].width = 100


def aba_resumo(planilha, rodadas):
    linhas = []

    for rodada in rodadas:
        t = rodada["totais"]
        meta = rodada["metadados"]
        tempos = rodada["tempos"]

        linhas.append(
            [
                rodada["checkpoint"],
                meta.get("epocas", ""),
                meta.get("mAP@[.5:.95]", ""),
                meta.get("loss_final", ""),
                rodada["imagens"],
                t["anotadas"],
                t["previstas"],
                t["vp"],
                t["trocas"],
                t["fp"],
                t["fn"],
                round(t["precisao"], 3),
                round(t["revocacao"], 3),
                round(t["f1"], 3),
                round(t["localizadas"], 3),
                round(sum(tempos) / len(tempos)) if tempos else "",
                rodada["tags_lidas"],
            ]
        )

    escrever(
        planilha.create_sheet("Resumo"),
        [
            "Checkpoint",
            "Épocas",
            "mAP treino",
            "Loss final",
            "Imagens",
            "Anotadas",
            "Previstas",
            "VP",
            "Classe trocada",
            "FP",
            "FN",
            "Precisão",
            "Revocação",
            "F1",
            "Localizadas",
            "Tempo médio (ms)",
            "TAGs lidas",
        ],
        linhas,
        larguras=[42, 8, 11, 11, 9, 10, 10, 7, 15, 7, 7, 10, 11, 8, 12, 17, 11],
    )


def aba_curva(planilha, rodadas, indice_gabarito, limiar_iou):
    linhas = []

    for rodada in rodadas:
        for limiar in LIMIARES_CURVA:
            matriz, _, _, _ = avaliar(
                rodada["ids"], indice_gabarito, limiar_iou, limiar
            )
            t = totais(matriz)

            linhas.append(
                [
                    rodada["checkpoint"],
                    limiar,
                    t["previstas"],
                    t["vp"],
                    t["trocas"],
                    t["fp"],
                    t["fn"],
                    round(t["precisao"], 3),
                    round(t["revocacao"], 3),
                    round(t["f1"], 3),
                    round(t["localizadas"], 3),
                ]
            )

    escrever(
        planilha.create_sheet("Curva"),
        [
            "Checkpoint",
            "Limiar de score",
            "Previstas",
            "VP",
            "Classe trocada",
            "FP",
            "FN",
            "Precisão",
            "Revocação",
            "F1",
            "Localizadas",
        ],
        linhas,
        larguras=[42, 15, 11, 7, 15, 7, 7, 10, 11, 8, 12],
    )


def aba_por_classe(planilha, rodadas, treino):
    linhas = []

    for rodada in rodadas:
        matriz = rodada["matriz"]

        classes = sorted(
            {r for r, _ in matriz if r != SEM_PAR}
            | {p for _, p in matriz if p != SEM_PAR}
            | set(treino)
        )

        for classe in classes:
            anotadas = sum(q for (r, _), q in matriz.items() if r == classe)
            previstas = sum(q for (_, p), q in matriz.items() if p == classe)
            acertos = matriz.get((classe, classe), 0)

            linhas.append(
                [
                    rodada["checkpoint"],
                    classe,
                    treino.get(classe, 0),
                    anotadas,
                    previstas,
                    acertos,
                    previstas - acertos,
                    anotadas - acertos,
                    round(acertos / previstas, 3) if previstas else 0.0,
                    round(acertos / anotadas, 3) if anotadas else 0.0,
                    _diagnostico(treino.get(classe, 0), anotadas, acertos),
                ]
            )

    escrever(
        planilha.create_sheet("Por classe"),
        [
            "Checkpoint",
            "Classe",
            "Exemplos no treino",
            "Anotadas no teste",
            "Previstas",
            "VP",
            "FP",
            "FN",
            "Precisão",
            "Revocação",
            "Leitura",
        ],
        linhas,
        larguras=[42, 22, 18, 18, 11, 7, 7, 7, 10, 11, 46],
    )


# Abaixo disto, a revocação da classe é ruído amostral: 0 de 4 não
# distingue um modelo cego de azar. Foi um erro que já se cometeu ao ler
# este relatório — "Conexão" foi apontada como anomalia com base em 66
# exemplos de treino e revocação zero, quando o teste tinha 4 anotações
# numa única imagem.
SUPORTE_MINIMO = 10


def _diagnostico(exemplos_treino: int, anotadas: int, acertos: int) -> str:
    """
    Leitura sugerida de cada linha, para não confundir três problemas
    diferentes que produzem o mesmo zero na revocação.
    """
    if anotadas == 0:
        return "classe ausente do conjunto de teste — nada a concluir"

    if anotadas < SUPORTE_MINIMO:
        return (
            f"amostra pequena demais ({anotadas} no teste) — "
            f"a revocação aqui é ruído, não conclua nada"
        )

    revocacao = acertos / anotadas

    if revocacao >= 0.4:
        return "desempenho utilizável"

    if exemplos_treino < 15:
        return (
            f"poucos exemplos no treino ({exemplos_treino}) — "
            f"falta dado, épocas não resolvem"
        )

    if revocacao < 0.1:
        return (
            f"{exemplos_treino} exemplos de treino e revocação ~0 — "
            f"suspeitar da anotação antes de treinar mais"
        )

    return "subtreino provável — mais épocas devem ajudar"


def aba_por_imagem(planilha, rodadas):
    linhas = []

    for rodada in rodadas:
        for item in sorted(rodada["por_imagem"], key=lambda i: i["revocacao"]):
            linhas.append(
                [
                    rodada["checkpoint"],
                    item["arquivo"],
                    item["dimensoes"],
                    item["anotadas"],
                    item["previstas"],
                    item["vp"],
                    item["trocas"],
                    item["fp"],
                    item["fn"],
                    round(item["revocacao"], 3),
                    item["tags_lidas"],
                    round(item["tempo_total_ms"]) if item["tempo_total_ms"] else "",
                ]
            )

    escrever(
        planilha.create_sheet("Por imagem"),
        [
            "Checkpoint",
            "Arquivo",
            "Dimensões",
            "Anotadas",
            "Previstas",
            "VP",
            "Classe trocada",
            "FP",
            "FN",
            "Revocação",
            "TAGs lidas",
            "Tempo (ms)",
        ],
        linhas,
        larguras=[42, 14, 12, 11, 11, 7, 15, 7, 7, 11, 11, 11],
    )


def aba_confusoes(planilha, rodadas):
    linhas = []

    for rodada in rodadas:
        confusoes = sorted(
            (
                (q, r, p)
                for (r, p), q in rodada["matriz"].items()
                if r != p and r != SEM_PAR and p != SEM_PAR
            ),
            reverse=True,
        )

        for quantidade, real, previsto in confusoes:
            linhas.append([rodada["checkpoint"], real, previsto, quantidade])

    escrever(
        planilha.create_sheet("Confusões"),
        ["Checkpoint", "Classe real", "Classe prevista", "Ocorrências"],
        linhas,
        larguras=[42, 22, 22, 13],
    )


def aba_deteccoes(planilha, rodadas):
    ordem = {"FN": 0, "CLASSE TROCADA": 1, "FP": 2, "VP": 3}

    linhas = []

    for rodada in rodadas:
        detalhes = sorted(
            rodada["detalhes"],
            key=lambda d: (ordem.get(d["situacao"], 9), d["arquivo"] or ""),
        )

        for d in detalhes:
            linhas.append(
                [
                    rodada["checkpoint"],
                    d["arquivo"],
                    d["situacao"],
                    d["classe_real"],
                    d["classe_prevista"],
                    d["score"],
                    d["iou"],
                    d["x1"],
                    d["y1"],
                    d["x2"],
                    d["y2"],
                    d["tag"],
                    d["texto_ocr"],
                ]
            )

    escrever(
        planilha.create_sheet("Detecções"),
        [
            "Checkpoint",
            "Arquivo",
            "Situação",
            "Classe real",
            "Classe prevista",
            "Score",
            "IoU",
            "x1",
            "y1",
            "x2",
            "y2",
            "TAG",
            "Texto do OCR",
        ],
        linhas,
        larguras=[42, 14, 16, 20, 20, 9, 8, 7, 7, 7, 7, 14, 16],
    )


# ======================================================================
# CLI
# ======================================================================


def montar_rodadas(ids, indice_gabarito, limiar_iou, limiar_score):
    """Agrupa as execuções por checkpoint — cada grupo é uma rodada."""
    por_checkpoint = defaultdict(list)

    for resumo in execucoes.listar(limite=100000):
        if resumo.id in ids:
            por_checkpoint[resumo.checkpoint or "(sem checkpoint)"].append(
                resumo
            )

    rodadas = []

    for checkpoint, lista in por_checkpoint.items():
        execucao_ids = [r.id for r in lista]

        matriz, por_imagem, detalhes, tempos = avaliar(
            execucao_ids, indice_gabarito, limiar_iou, limiar_score
        )

        if not por_imagem:
            continue

        rodadas.append(
            {
                "checkpoint": checkpoint,
                "ids": execucao_ids,
                "imagens": len(por_imagem),
                "matriz": matriz,
                "por_imagem": por_imagem,
                "detalhes": detalhes,
                "tempos": tempos,
                "totais": totais(matriz),
                "tags_lidas": sum(r.qtd_tags_lidas for r in lista),
                "metadados": metadados_checkpoint(checkpoint),
            }
        )

    return sorted(rodadas, key=lambda r: r["checkpoint"])


def parse_args():
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--saida", default="data/output/relatorio.xlsx")
    parser.add_argument("--gabarito", default="dataset/original")
    parser.add_argument("--treino", default="dataset/original/treinamento")
    parser.add_argument("--execucoes", help="ids separados por vírgula")
    parser.add_argument("--iou", type=float, default=0.5)
    parser.add_argument(
        "--score",
        type=float,
        default=0.05,
        help="Limiar usado nas abas de detalhe (a Curva varre todos).",
    )
    return parser.parse_args()


def main() -> None:
    forcar_utf8()

    args = parse_args()

    indice = gabarito.indexar(Path(args.gabarito))

    if not indice:
        raise SystemExit(f"Nenhum .xml de gabarito em {args.gabarito}")

    if args.execucoes:
        ids = {int(x) for x in args.execucoes.split(",") if x.strip()}
    else:
        ids = {e.id for e in execucoes.listar(limite=100000)}

    if not ids:
        raise SystemExit(
            "Nenhuma execução no banco. Rode "
            "'python -m app.diagnostico.lote' antes."
        )

    print("apurando...")
    rodadas = montar_rodadas(ids, indice, args.iou, args.score)

    if not rodadas:
        raise SystemExit(
            "Nenhuma execução tinha gabarito correspondente em "
            f"{args.gabarito}."
        )

    treino = exemplos_de_treino(args.treino)

    planilha = Workbook()

    aba_leiame(planilha, rodadas, args.iou, args.score)
    aba_resumo(planilha, rodadas)
    print("curva...")
    aba_curva(planilha, rodadas, indice, args.iou)
    aba_por_classe(planilha, rodadas, treino)
    aba_por_imagem(planilha, rodadas)
    aba_confusoes(planilha, rodadas)
    aba_deteccoes(planilha, rodadas)

    destino = Path(args.saida)
    destino.parent.mkdir(parents=True, exist_ok=True)
    planilha.save(destino)

    print()
    for rodada in rodadas:
        t = rodada["totais"]
        print(
            f"  {rodada['checkpoint']}\n"
            f"    {rodada['imagens']} imagens | VP={t['vp']} FP={t['fp']} "
            f"FN={t['fn']} troca={t['trocas']} | "
            f"precisão={t['precisao']:.3f} revocação={t['revocacao']:.3f}"
        )

    print()
    print(f"Relatório salvo em: {destino.resolve()}")


if __name__ == "__main__":
    main()
