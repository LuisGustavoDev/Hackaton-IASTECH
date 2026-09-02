"""
Integração: imagem real de P&ID -> planilha.

Percorre o caminho inteiro (validação -> detecção -> OCR -> associação ->
ISA -> planilha) usando imagens do próprio dataset do projeto.

O checkpoint é montado na hora com pesos aleatórios, a menos que a máquina
tenha um modelo treinado em DETECTOR_CHECKPOINT_PATH. Com pesos aleatórios
as detecções não significam nada — o que está sendo verificado é que o
encanamento inteiro roda de ponta a ponta e produz a planilha no formato
combinado.
"""

from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

torch = pytest.importorskip("torch")
pytest.importorskip("torchvision")

from app.core.erros import ImagemInvalidaError  # noqa: E402
from app.detection.checkpoint import salvar_checkpoint  # noqa: E402
from app.detection.detector import redefinir_detector  # noqa: E402
from app.detection.modelo import construir_faster_rcnn  # noqa: E402
from app.services import image_service, processamento  # noqa: E402
from app.services.planilha_service import COLUNAS  # noqa: E402

pytestmark = pytest.mark.lento

# As 24 classes de equipamento do dataset anotado do projeto.
CLASSES = [
    "Acumulador", "Agitador", "Bomba", "Caldeira", "Compressor", "Conexão",
    "Filtro", "Forno", "HVAC", "Instrumento", "Misturador", "Motor",
    "Outro", "Reator", "Resfriador", "Separador", "SistemadePotência",
    "Tanque", "TrocadordeCalor", "Tubulação", "Turbina", "Válvula",
    "Vaso", "Ventilador",
]


@pytest.fixture(scope="module")
def checkpoint(tmp_path_factory) -> Path:
    treinado = os.environ.get("DETECTOR_CHECKPOINT_PATH")

    if treinado and Path(treinado).is_file():
        return Path(treinado)

    destino = tmp_path_factory.mktemp("modelos") / "faster_rcnn.pt"
    modelo = construir_faster_rcnn(len(CLASSES), pesos_pretreinados=False)

    return salvar_checkpoint(
        destino, modelo, CLASSES, metadados={"origem": "teste de integração"}
    )


@pytest.fixture
def pipeline_configurado(checkpoint, monkeypatch, tmp_path):
    monkeypatch.setenv("DETECTOR_CHECKPOINT_PATH", str(checkpoint))
    # Isola os artefatos do teste da pasta data/output do projeto.
    monkeypatch.setattr(processamento, "DIRETORIO_SAIDA", tmp_path / "saida")
    redefinir_detector()


@pytest.fixture
def ocr_simulado(monkeypatch):
    """
    Substitui o OCR por tags conhecidas.

    Serve para que o teste rode em máquinas sem o binário do Tesseract
    (ele só existe dentro da imagem Docker) e para que as colunas
    Descrição e Grupo tenham entrada previsível.
    """
    def textos_fixos(_caminho):
        return [
            {"id": 1, "texto": "FT-210", "confianca": 92.0,
             "x": 100, "y": 100, "w": 50, "h": 14, "crop": None},
            {"id": 2, "texto": "PI-101", "confianca": 88.0,
             "x": 300, "y": 250, "w": 50, "h": 14, "crop": None},
        ]

    monkeypatch.setattr(image_service, "processar_imagem", textos_fixos)


def test_imagem_real_gera_planilha_no_formato_combinado(
    imagem_real, pipeline_configurado, ocr_simulado
):
    resultado = processamento.processar_imagem(imagem_real.read_bytes())

    assert resultado.planilha_xlsx.is_file()
    assert resultado.planilha_csv.is_file()

    aba = load_workbook(resultado.planilha_xlsx).active

    assert [celula.value for celula in aba[1]] == COLUNAS


def test_imagem_real_produz_linhas_consistentes(
    imagem_real, pipeline_configurado, ocr_simulado
):
    resultado = processamento.processar_imagem(imagem_real.read_bytes())

    for linha in resultado.linhas:
        assert set(linha) == set(COLUNAS)
        assert linha["Tipo"] in CLASSES
        # Descrição nunca fica vazia: sem TAG, cai para a classe detectada.
        assert linha["Descrição"]
        assert isinstance(linha["Coordenada X"], int)
        assert isinstance(linha["Coordenada Y"], int)


def test_coordenadas_ficam_dentro_da_imagem(
    imagem_real, pipeline_configurado, ocr_simulado
):
    import cv2

    altura, largura = cv2.imread(str(imagem_real)).shape[:2]

    resultado = processamento.processar_imagem(imagem_real.read_bytes())

    for linha in resultado.linhas:
        assert 0 <= linha["Coordenada X"] <= largura
        assert 0 <= linha["Coordenada Y"] <= altura


def test_tag_lida_preenche_descricao_e_grupo(
    imagem_real, pipeline_configurado, ocr_simulado
):
    """
    As duas tags do OCR simulado ficam dentro de alguma bounding box em
    quase toda execução; quando ficam, Descrição e Grupo têm de vir da
    decomposição ISA e não da classe detectada.
    """
    resultado = processamento.processar_imagem(imagem_real.read_bytes())

    com_tag = [linha for linha in resultado.linhas if linha["TAG"]]

    for linha in com_tag:
        assert linha["Grupo"] in {"1", "2"}
        assert linha["Descrição"] in {
            "Transmissor de Vazão",
            "Indicador de Pressão",
        }


def test_arquivo_que_nao_e_imagem_para_antes_do_modelo(pipeline_configurado):
    """
    A validação roda antes de qualquer coisa cara: um PDF renomeado é
    recusado sem sequer carregar o detector.
    """
    with pytest.raises(ImagemInvalidaError, match="não é uma imagem"):
        processamento.processar_imagem(b"%PDF-1.7\n%fake\n")


def test_imagem_corrompida_para_antes_do_modelo(
    imagem_real, pipeline_configurado
):
    conteudo = imagem_real.read_bytes()

    with pytest.raises(ImagemInvalidaError):
        processamento.processar_imagem(conteudo[: len(conteudo) // 3])


def test_pipeline_completo_com_ocr_real(
    imagem_real, pipeline_configurado, tesseract_disponivel
):
    """
    Sem simulação nenhuma: detector + Tesseract de verdade sobre uma
    imagem do dataset. Pulado onde o binário do Tesseract não existe (fora
    do container).
    """
    resultado = processamento.processar_imagem(imagem_real.read_bytes())

    assert resultado.planilha_xlsx.is_file()
    assert load_workbook(resultado.planilha_xlsx).active.max_row >= 1
