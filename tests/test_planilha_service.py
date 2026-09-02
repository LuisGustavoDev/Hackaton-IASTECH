"""
Exportação da planilha.

As colunas e a ordem foram fechadas com o cliente — este teste existe para
que uma mudança acidental nelas quebre o build.
"""

from __future__ import annotations

import csv

from openpyxl import load_workbook

from app.services.planilha_service import COLUNAS, gerar_planilha

LINHAS = [
    {
        "TAG": "FT-210",
        "Tipo": "Instrumento",
        "Descrição": "Transmissor de Vazão",
        "Coordenada X": 512,
        "Coordenada Y": 340,
        "Grupo": "2",
    },
    {
        "TAG": "",
        "Tipo": "Válvula",
        "Descrição": "Válvula",
        "Coordenada X": 88,
        "Coordenada Y": 91,
        "Grupo": "",
    },
]


def test_colunas_sao_exatamente_as_combinadas():
    assert COLUNAS == [
        "TAG",
        "Tipo",
        "Descrição",
        "Coordenada X",
        "Coordenada Y",
        "Grupo",
    ]


def test_xlsx_tem_o_cabecalho_e_os_dados(tmp_path):
    xlsx, _ = gerar_planilha(
        LINHAS, tmp_path / "r.xlsx", tmp_path / "r.csv"
    )

    aba = load_workbook(xlsx).active

    assert [c.value for c in aba[1]] == COLUNAS
    assert [c.value for c in aba[2]] == [
        "FT-210",
        "Instrumento",
        "Transmissor de Vazão",
        512,
        340,
        "2",
    ]
    assert aba.max_row == 3


def test_csv_tem_o_cabecalho_e_os_dados(tmp_path):
    _, arquivo_csv = gerar_planilha(
        LINHAS, tmp_path / "r.xlsx", tmp_path / "r.csv"
    )

    with arquivo_csv.open(encoding="utf-8-sig") as arquivo:
        linhas = list(csv.reader(arquivo))

    assert linhas[0] == COLUNAS
    assert linhas[1][0] == "FT-210"
    assert linhas[2][0] == ""


def test_acentos_sobrevivem_ao_csv(tmp_path):
    _, arquivo_csv = gerar_planilha(
        [
            {
                "TAG": "V-1",
                "Tipo": "Conexão",
                "Descrição": "Válvula",
                "Coordenada X": 1,
                "Coordenada Y": 2,
                "Grupo": "1",
            }
        ],
        tmp_path / "r.xlsx",
        tmp_path / "r.csv",
    )

    conteudo = arquivo_csv.read_text(encoding="utf-8-sig")

    assert "Conexão" in conteudo
    assert "Válvula" in conteudo


def test_planilha_vazia_ainda_tem_cabecalho(tmp_path):
    xlsx, _ = gerar_planilha([], tmp_path / "r.xlsx", tmp_path / "r.csv")

    aba = load_workbook(xlsx).active

    assert [c.value for c in aba[1]] == COLUNAS
    assert aba.max_row == 1


def test_cria_o_diretorio_de_destino(tmp_path):
    destino = tmp_path / "nao" / "existe"

    xlsx, arquivo_csv = gerar_planilha(
        LINHAS, destino / "r.xlsx", destino / "r.csv"
    )

    assert xlsx.is_file()
    assert arquivo_csv.is_file()
