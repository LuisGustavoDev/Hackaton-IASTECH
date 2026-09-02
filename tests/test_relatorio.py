"""
Relatório comparativo em Excel.
"""

from __future__ import annotations

import pytest
from collections import Counter
from openpyxl import Workbook, load_workbook

from app.diagnostico import gabarito
from app.diagnostico.matriz_confusao import SEM_PAR, casar, casar_detalhado
from app.diagnostico.relatorio import (
    aba_por_classe,
    exemplos_de_treino,
    metadados_checkpoint,
    totais,
)


def caixa(classe, x1, y1, x2, y2, score=0.9):
    return {
        "classe": classe,
        "score": score,
        "x1": x1,
        "y1": y1,
        "x2": x2,
        "y2": y2,
    }


# ---------------------------------------------------------------------
# Casamento detalhado
# ---------------------------------------------------------------------


def test_casar_detalhado_preserva_os_objetos_e_o_iou():
    anotacao = caixa("Válvula", 0, 0, 10, 10)
    predicao = caixa("Válvula", 0, 0, 10, 10, score=0.8)

    pares, fp, fn = casar_detalhado([predicao], [anotacao], 0.5)

    (par_anotacao, par_predicao, valor_iou) = pares[0]

    assert par_anotacao is anotacao
    assert par_predicao is predicao
    assert valor_iou == pytest.approx(1.0)
    assert fp == [] and fn == []


def test_casar_resumido_continua_devolvendo_nomes():
    """A versão resumida é usada pela matriz; não pode ter mudado."""
    pares, fp, fn = casar(
        [caixa("Outro", 0, 0, 10, 10)],
        [caixa("Válvula", 0, 0, 10, 10)],
        0.5,
    )

    assert pares == [("Válvula", "Outro")]
    assert fp == [] and fn == []


def test_casar_detalhado_devolve_os_nao_casados_como_objetos():
    predicao = caixa("Bomba", 0, 0, 10, 10)
    anotacao = caixa("Vaso", 500, 500, 510, 510)

    pares, fp, fn = casar_detalhado([predicao], [anotacao], 0.5)

    assert pares == []
    assert fp == [predicao]
    assert fn == [anotacao]


# ---------------------------------------------------------------------
# Totais
# ---------------------------------------------------------------------


def test_totais_separa_troca_de_falso_positivo():
    matriz = Counter(
        {
            ("Válvula", "Válvula"): 10,
            ("Válvula", "Outro"): 4,
            (SEM_PAR, "Outro"): 3,
            ("Tanque", SEM_PAR): 2,
        }
    )

    t = totais(matriz)

    assert t["vp"] == 10
    assert t["trocas"] == 4
    assert t["fp"] == 3
    assert t["fn"] == 2
    assert t["anotadas"] == 16
    assert t["previstas"] == 17


def test_localizadas_conta_a_caixa_certa_mesmo_com_rotulo_errado():
    """
    "Localizadas" mede quanto o modelo ACHOU, independente do rótulo — é
    o número que separa problema de detecção de problema de classificação.
    """
    t = totais(
        Counter({("Válvula", "Válvula"): 3, ("Válvula", "Outro"): 5, ("Vaso", SEM_PAR): 2})
    )

    assert t["anotadas"] == 10
    assert t["localizadas"] == pytest.approx(0.8)
    assert t["revocacao"] == pytest.approx(0.3)


def test_totais_de_matriz_vazia_nao_divide_por_zero():
    t = totais(Counter())

    assert t["precisao"] == 0.0
    assert t["revocacao"] == 0.0
    assert t["f1"] == 0.0
    assert t["localizadas"] == 0.0


# ---------------------------------------------------------------------
# Exemplos de treino
# ---------------------------------------------------------------------


def test_exemplos_de_treino_conta_por_classe():
    contagem = exemplos_de_treino("dataset/original/treinamento")

    if not contagem:
        pytest.skip("dataset não presente nesta cópia do repositório")

    assert contagem["Válvula"] > contagem["Acumulador"]
    assert sum(contagem.values()) > 1000


def test_metadados_de_checkpoint_inexistente_nao_quebra(tmp_path):
    """
    O relatório precisa funcionar numa máquina que tem o banco mas não
    tem mais o .pt daquela rodada.
    """
    assert metadados_checkpoint(str(tmp_path / "nao_existe.pt")) == {}


def test_metadados_de_arquivo_invalido_nao_quebra(tmp_path):
    caminho = tmp_path / "lixo.pt"
    caminho.write_bytes(b"nao e um checkpoint")

    assert metadados_checkpoint(str(caminho)) == {}


# ---------------------------------------------------------------------
# Abas
# ---------------------------------------------------------------------


def test_por_classe_cruza_desempenho_com_exemplos_de_treino(tmp_path):
    """
    A coluna "exemplos no treino" ao lado da revocação é o que separa
    subtreino (muitos exemplos, revocação baixa) de falta de dado
    (poucos exemplos, revocação zero). Sem ela a tabela não diagnostica
    nada.
    """
    planilha = Workbook()

    rodada = {
        "checkpoint": "modelo.pt",
        "matriz": Counter(
            {("Válvula", "Válvula"): 8, ("Vaso", SEM_PAR): 3}
        ),
    }

    aba_por_classe(planilha, [rodada], Counter({"Válvula": 728, "Vaso": 11}))

    destino = tmp_path / "r.xlsx"
    planilha.save(destino)

    aba = load_workbook(destino)["Por classe"]

    cabecalho = [c.value for c in aba[1]]
    assert "Exemplos no treino" in cabecalho

    coluna_treino = cabecalho.index("Exemplos no treino") + 1
    coluna_classe = cabecalho.index("Classe") + 1

    por_classe = {
        aba.cell(row=r, column=coluna_classe).value: aba.cell(
            row=r, column=coluna_treino
        ).value
        for r in range(2, aba.max_row + 1)
    }

    assert por_classe["Válvula"] == 728
    assert por_classe["Vaso"] == 11


def test_iou_entra_no_relatorio_com_o_valor_do_casamento():
    """Interseção 5x10=50, união 150 -> 1/3."""
    anotacao = caixa("Válvula", 0, 0, 10, 10)
    predicao = caixa("Válvula", 5, 0, 15, 10)

    assert gabarito.iou(anotacao, predicao) == pytest.approx(1 / 3)
